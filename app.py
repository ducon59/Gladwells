import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import msoffcrypto
import openpyxl
import io
import os
import re
import requests
from datetime import datetime
import time

# ─── Config ───────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Gladwells Management Accounts",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

FILE_PASSWORD   = "Dunnock057"
# Folder ID baked in via Streamlit secrets or env var.
# The app always picks the most-recently-modified .xlsm in this folder.
GDRIVE_FOLDER_ID = (
    st.secrets.get("GDRIVE_FOLDER_ID", "")
    if hasattr(st, "secrets")
    else os.environ.get("GDRIVE_FOLDER_ID", "")
)

# ─── Colours ──────────────────────────────────────────────────────────────────
PRIMARY   = "#2C5F2E"
ACCENT    = "#97BC62"
ORANGE    = "#E8623A"
BG        = "#F5F7F2"
CARD_BG   = "#FFFFFF"
TEXT_DARK = "#1A2E1A"

# ─── CSS ──────────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
  @import url('https://fonts.googleapis.com/css2?family=Raleway:wght@700;800;900&display=swap');

  html, body, [class*="css"] {{ font-family: 'Inter', sans-serif; background: {BG}; }}
  .main {{ background: {BG}; }}
  .stSidebar {{ background: {PRIMARY}; }}
  .stSidebar * {{ color: white !important; }}

  /* ── Big title ── */
  .dashboard-title {{
    font-family: 'Raleway', sans-serif;
    font-size: 48px;
    font-weight: 900;
    color: {PRIMARY};
    letter-spacing: -1px;
    line-height: 1.1;
    margin: 0 0 4px 0;
  }}
  .dashboard-subtitle {{
    font-size: 15px;
    color: #777;
    margin: 0;
    font-weight: 400;
  }}

  /* ── Tabs ── */
  .stTabs [data-baseweb="tab-list"] {{
    gap: 0px;
    background: {PRIMARY};
    border-radius: 10px 10px 0 0;
    padding: 0 8px;
    width: 100%;
  }}
  .stTabs [data-baseweb="tab"] {{
    flex: 1;
    justify-content: center;
    height: 54px;
    background: transparent;
    border-radius: 0;
    color: rgba(255,255,255,0.65) !important;
    font-size: 15px;
    font-weight: 600;
    letter-spacing: 0.3px;
    border-bottom: 3px solid transparent;
    transition: all 0.2s;
  }}
  .stTabs [aria-selected="true"] {{
    background: rgba(255,255,255,0.12) !important;
    color: white !important;
    border-bottom: 3px solid {ACCENT} !important;
  }}
  .stTabs [data-baseweb="tab"]:hover {{
    color: white !important;
    background: rgba(255,255,255,0.08) !important;
  }}
  .stTabs [data-baseweb="tab-panel"] {{
    background: white;
    border-radius: 0 0 10px 10px;
    padding: 28px 24px;
    box-shadow: 0 4px 16px rgba(0,0,0,0.08);
  }}

  /* ── Metric cards ── */
  .metric-card {{
    background: {CARD_BG};
    border-radius: 10px;
    padding: 16px 18px;
    border-left: 4px solid {ACCENT};
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    margin-bottom: 6px;
    height: 100%;
  }}
  .metric-card.prev-year {{
    border-left-color: #aaa;
    background: #fafafa;
  }}
  .metric-title {{
    font-size: 11px; font-weight: 600; color: #999;
    text-transform: uppercase; letter-spacing: 0.6px; margin-bottom: 4px;
  }}
  .metric-value {{ font-size: 26px; font-weight: 700; color: {TEXT_DARK}; line-height: 1.2; }}
  .metric-delta {{ font-size: 12px; font-weight: 500; margin-top: 4px; }}
  .delta-pos {{ color: #2a8a2e; }}
  .delta-neg {{ color: {ORANGE}; }}
  .row-label {{
    font-size: 11px; font-weight: 700; color: #555;
    text-transform: uppercase; letter-spacing: 0.8px;
    padding: 10px 0 6px 2px;
  }}

  /* ── Section headers ── */
  .section-header {{
    font-size: 17px; font-weight: 700; color: {TEXT_DARK};
    border-bottom: 2px solid {ACCENT}; padding-bottom: 8px;
    margin: 28px 0 16px 0;
  }}

  /* ── Legend hint ── */
  .legend-hint {{
    font-size: 12px; color: #999; font-style: italic;
    text-align: right; margin-top: -8px; margin-bottom: 4px;
  }}
</style>
""", unsafe_allow_html=True)

# ─── Data loading ─────────────────────────────────────────────────────────────
def _gdrive_download_file(file_id: str) -> bytes:
    """Download a single Drive file by ID, handling the large-file confirmation redirect."""
    session = requests.Session()
    base_url = "https://drive.google.com/uc"
    params = {"id": file_id, "export": "download"}
    response = session.get(base_url, params=params, stream=True, timeout=60)
    response.raise_for_status()
    if "text/html" in response.headers.get("Content-Type", ""):
        token_match = re.search(r'confirm=([0-9A-Za-z_-]+)', response.text)
        params["confirm"] = token_match.group(1) if token_match else "t"
        response = session.get(base_url, params=params, stream=True, timeout=60)
        response.raise_for_status()
    buf = io.BytesIO()
    for chunk in response.iter_content(chunk_size=32768):
        if chunk: buf.write(chunk)
    return buf.getvalue()




def _gdrive_folder_latest_file(folder_id_or_sheet_csv_url: str) -> tuple[str, str]:
    """
    Resolve a folder/pointer config to (file_id, filename).

    Supports two input formats:
      A) A Google Sheets CSV URL  (recommended, most reliable)
         Publish a Google Sheet as CSV; put the .xlsm file ID in cell A1
         and the filename in cell B1.  Update A1 each month.
         URL format: https://docs.google.com/spreadsheets/d/SHEET_ID/export?format=csv&gid=0

      B) A plain Google Drive file ID  (user pastes the .xlsm file ID directly)
         Falls back to treating the value as a direct file ID with a generic filename.
    """
    val = folder_id_or_sheet_csv_url.strip()

    # ── A: Google Sheets CSV pointer ──────────────────────────────────────────
    is_sheets_url = "docs.google.com/spreadsheets" in val or "export?format=csv" in val
    # Also handle the short published-CSV URL pattern
    is_csv_url = val.startswith("http") and ("format=csv" in val or "output=csv" in val)

    if is_sheets_url or is_csv_url:
        try:
            r = requests.get(val, timeout=15, allow_redirects=True)
            r.raise_for_status()
            text = r.text.strip()
            # Parse first row: file_id, filename
            first_row = text.splitlines()[0] if text else ""
            parts = [p.strip().strip('"') for p in first_row.split(",")]
            if not parts or not parts[0]:
                raise RuntimeError("Index sheet is empty — put the file ID in cell A1.")
            file_id = parts[0]
            filename = parts[1] if len(parts) > 1 and parts[1] else "management_accounts.xlsm"
            return file_id, filename
        except requests.RequestException as e:
            raise RuntimeError(
                f"Could not read the index sheet: {e}\n"
                "Make sure the Google Sheet is published: "
                "File → Share → Publish to web → CSV."
            ) from e

    # ── B: Bare file ID (direct) ──────────────────────────────────────────────
    # If the user pasted a Drive file ID directly, just use it
    if re.match(r'^[A-Za-z0-9_-]{25,60}$', val):
        return val, "management_accounts.xlsm"

    # ── C: Full Drive file URL ─────────────────────────────────────────────────
    m = re.search(r'/file/d/([A-Za-z0-9_-]{25,60})', val)
    if m:
        return m.group(1), "management_accounts.xlsm"

    raise RuntimeError(
        "Could not resolve the Google Drive config.\n\n"
        "The simplest setup is:\n"
        "1. Create a Google Sheet with the .xlsm file ID in cell A1 and "
        "filename in B1\n"
        "2. File → Share → Publish to web → Sheet1 → CSV → Publish\n"
        "3. Copy that CSV URL into the GDRIVE_FOLDER_ID secret\n\n"
        "Alternatively paste a direct Drive file ID or file URL.\n"
        f"Value received: {val[:80]}"
    )

def load_latest_from_folder(folder_id: str) -> tuple:
    """
    Find the most-recently-modified spreadsheet in the folder,
    download and decrypt it. Returns (BytesIO_buf, filename, modified_time).
    """
    file_id, filename = _gdrive_folder_latest_file(folder_id)
    raw_bytes = _gdrive_download_file(file_id)
    office_file = msoffcrypto.OfficeFile(io.BytesIO(raw_bytes))
    office_file.load_key(password=FILE_PASSWORD)
    buf_out = io.BytesIO()
    office_file.decrypt(buf_out)
    buf_out.seek(0)
    return buf_out, filename

def decrypt_local(uploaded_file) -> io.BytesIO:
    office_file = msoffcrypto.OfficeFile(uploaded_file)
    office_file.load_key(password=FILE_PASSWORD)
    buf = io.BytesIO()
    office_file.decrypt(buf)
    buf.seek(0)
    return buf

@st.cache_data(ttl=3600)
def parse_workbook(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    data = {}

    # ── P&L data ──
    ws = wb["P&L_Data"]
    rows = list(ws.iter_rows(values_only=True))
    header_row = rows[4]
    months = [c for c in header_row[2:14] if c is not None]

    def get_row(label):
        for r in rows:
            if r[0] == label or (len(r) > 1 and r[1] == label):
                return r
        return None

    def row_values(label):
        r = get_row(label)
        if r is None: return [0] * len(months)
        vals = list(r[2:2+len(months)])
        return [v if isinstance(v, (int, float)) else 0 for v in vals]

    data["months"]       = list(months)
    data["revenue"]      = row_values("Total Income")
    data["cogs"]         = row_values("Total Less Cost of Sales")
    data["gross_profit"] = row_values("Gross Profit")
    data["total_opex"]   = row_values("Total Less Operating Expenses")
    data["op_profit"]    = row_values("Operating Profit")
    data["net_profit"]   = row_values("Net Profit")
    data["staff_costs"]  = row_values("Total 0 Staff Costs")
    data["rent_util"]    = row_values("Total 1 Rent, Rates, Utilities")
    data["maintenance"]  = row_values("Total 2 Shop Maintenance")
    data["other_oh"]     = row_values("Total 3 Other overheads")
    data["depreciation"] = row_values("Depreciation")

    income_lines = [
        "Coffee and Tea Sales", "Fruit and Veg Sales", "Grocery  Sales",
        "In-House Food Sales", "Turner & George sales",
        "Wine, Beer and Spirits Sales", "Homeware and Gifting Sales",
        "Deli Sales", "Cafe Sales", "Plants and Flowers sales"
    ]
    data["revenue_breakdown"] = {k: row_values(k) for k in income_lines}

    # ── Stats sheet ──
    ws2 = wb["Stats"]
    srows = list(ws2.iter_rows(min_row=1, values_only=True))
    stat_months, turnover, gp_pct, overheads, cash = [], [], [], [], []
    for r in srows:
        if r[0] == "Turnover":
            turnover = [v if isinstance(v, (int, float)) else 0 for v in r[1:13]]
        elif r[0] == "Gross profit":
            gp_pct = [v if isinstance(v, (int, float)) else 0 for v in r[1:13]]
        elif r[0] == "Overheads":
            overheads = [v if isinstance(v, (int, float)) else 0 for v in r[1:13]]
        elif r[0] == "Cash balance each month":
            cash = [v if isinstance(v, (int, float)) else 0 for v in r[1:13]]
        elif r[0] is None and isinstance(r[1], str) and r[1].startswith("Mar"):
            stat_months = list(r[1:13])
    data["stat_months"]   = stat_months
    data["stat_turnover"] = turnover
    data["stat_gp_pct"]   = gp_pct
    data["stat_overheads"]= overheads
    data["stat_cash"]     = cash

    # ── Manual / KPIs ──
    ws3 = wb["Manual"]
    mrows = list(ws3.iter_rows(values_only=True))
    for r in mrows:
        if len(r) > 4 and r[1] == "Number of Transactions p.w.":
            data["txn_curr"] = r[2]; data["txn_prev"] = r[3]; data["txn_prev_yr"] = r[4]
        elif len(r) > 4 and r[1] == "Average Spend Per Head (ex. VAT)":
            data["asp_curr"] = r[2]; data["asp_prev"] = r[3]; data["asp_prev_yr"] = r[4]

    # ── Overview ── col[1]=curr, col[2]=prev mth, col[5]=prev yr, col[8]=YTD, col[9]=YTD LY
    ws4 = wb["Overview"]
    orows = list(ws4.iter_rows(min_row=5, values_only=True))
    for r in orows:
        lbl = r[0]
        def _v(idx):
            v = r[idx] if len(r) > idx else None
            return v if isinstance(v, (int, float)) else None
        if lbl == "Sales":
            data["ov_sales_curr"]    = _v(1)
            data["ov_sales_prev_mth"]= _v(2)
            data["ov_sales_prev_yr"] = _v(5)
            data["ov_ytd"]           = _v(8)
            data["ov_ytd_ly"]        = _v(9)
        elif lbl == "Gross profit Total":
            data["ov_gp_curr"]     = _v(1)
            data["ov_gp_prev_mth"] = _v(2)
            data["ov_gp_prev_yr"]  = _v(5)
        elif lbl == "Gross Profit":
            data["ov_gp_pct_curr"]    = _v(1)
            data["ov_gp_pct_prev_mth"]= _v(2)
            data["ov_gp_pct_prev_yr"] = _v(5)
        elif lbl == "Staff Costs":
            data["ov_staff_curr"]    = _v(1)
            data["ov_staff_prev_mth"]= _v(2)
            data["ov_staff_prev_yr"] = _v(5)
        elif lbl == "Total Staff / Sales":
            data["ov_staff_pct_curr"]    = _v(1)
            data["ov_staff_pct_prev_mth"]= _v(2)
            data["ov_staff_pct_prev_yr"] = _v(5)
        elif lbl == "Overheads":
            data["ov_oh_curr"]    = _v(1)
            data["ov_oh_prev_yr"] = _v(5)

    # ── Period label ──
    prows = list(ws.iter_rows(max_row=5, values_only=True))
    data["period_label"] = prows[2][0] if prows[2][0] else "Management Accounts"

    # ── Prior year same month net profit (col 5 from P&L rows in overview) ──
    # Net profit not directly in Overview, derive: op_profit - depreciation
    # Use P&L_Data last column set: months[12] is prior year same month
    data["net_profit_prev_yr"] = row_values("Net Profit")[12] if len(row_values("Net Profit")) > 12 else None

    # ── BS data ──
    ws7 = wb["BS_Data"]
    brows = list(ws7.iter_rows(values_only=True))
    bs_header = brows[3]
    bs_months = [c for c in bs_header[2:14] if c is not None]
    data["bs_months"] = [str(m)[:10] if isinstance(m, datetime) else str(m) for m in bs_months]
    for r in brows:
        if r[0] == "Total Fixed Assets":
            data["bs_fixed"]     = [v if isinstance(v, (int, float)) else 0 for v in r[2:14]]
        elif r[0] == "Total Current Assets":
            data["bs_current"]   = [v if isinstance(v, (int, float)) else 0 for v in r[2:14]]
        elif r[0] == "Total Current Liabilities":
            data["bs_curr_liab"] = [v if isinstance(v, (int, float)) else 0 for v in r[2:14]]
        elif r[0] == "Net Assets":
            data["bs_net_assets"]= [v if isinstance(v, (int, float)) else 0 for v in r[2:14]]
        elif r[0] == "Cash":
            data["bs_cash_bal"]  = [v if isinstance(v, (int, float)) else 0 for v in r[2:14]]

    return data

# ─── Helpers ──────────────────────────────────────────────────────────────────
def fmt_k(v):
    if v is None: return "—"
    return f"£{v/1000:,.1f}k"

def fmt_pct(v):
    if v is None: return "—"
    return f"{v*100:.1f}%"

def delta_html(curr, prev, is_pct=False, label="vs prior month", invert=False):
    if curr is None or prev is None: return ""
    diff = curr - prev
    positive = diff > 0
    if invert: positive = not positive
    cls = "delta-pos" if positive else "delta-neg"
    sign = "▲" if diff > 0 else "▼"
    if is_pct:
        return f'<span class="{cls}">{sign} {abs(diff)*100:.1f}pp {label}</span>'
    else:
        return f'<span class="{cls}">{sign} {fmt_k(abs(diff))} {label}</span>'

def legend_hint():
    st.markdown('<div class="legend-hint">💡 Click legend items to show/hide</div>', unsafe_allow_html=True)

def plotly_theme(legend_y=1.08):
    return dict(
        paper_bgcolor="white", plot_bgcolor="white",
        font=dict(family="Inter, sans-serif", size=12, color=TEXT_DARK),
        margin=dict(l=40, r=20, t=55, b=40),
        legend=dict(
            orientation="h", y=legend_y, x=0,
            bgcolor="rgba(255,255,255,0.9)",
            bordercolor="#ddd", borderwidth=1,
            font=dict(size=12),
            itemclick="toggle", itemdoubleclick="toggleothers",
            title=dict(text="<b>Click to toggle →</b>  ", font=dict(size=11, color="#888")),
        )
    )

# ─── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📊 Gladwells")
    st.markdown("**Management Accounts**")
    st.divider()

    # Admin override: allow pasting config in case secrets aren't set
    if not GDRIVE_FOLDER_ID:
        manual_folder = st.text_input(
            "Google Drive config",
            placeholder="Paste index sheet CSV URL or file ID…",
            help=(
                "Recommended: paste the CSV publish URL from your index Google Sheet. "
                "Or paste a direct Drive file ID. "
                "Add GDRIVE_FOLDER_ID to Streamlit secrets to remove this prompt permanently."
            )
        )
    else:
        manual_folder = ""
        st.success("✅ Connected to Google Drive", icon="☁️")

    if st.button("🔄 Refresh data", use_container_width=True):
        st.cache_data.clear()

    st.divider()
    st.markdown("**Manual upload**")
    uploaded = st.file_uploader(
        "Upload a specific file",
        type=["xlsm", "xlsx"],
        help="Optional: upload a file directly — overrides the Drive config."
    )
    st.divider()
    st.caption("Monthly update: update cell A1 of your index sheet with the new file's ID.")

# ─── Load data ────────────────────────────────────────────────────────────────
buf       = None
file_label = ""
folder_id = GDRIVE_FOLDER_ID or manual_folder

if uploaded:
    # Manual upload always takes priority
    buf = decrypt_local(uploaded)
    file_label = uploaded.name
elif folder_id:
    try:
        with st.spinner("Fetching latest file from Google Drive…"):
            buf, file_label = load_latest_from_folder(folder_id)
        with st.sidebar:
            st.info(f"📄 **{file_label}**", icon="📥")
    except Exception as e:
        st.error(f"Could not load from Google Drive folder: {e}")

if buf is None:
    folder_configured = bool(folder_id)
    st.markdown(f"""
    <div style="text-align:center; padding: 60px 20px; color: #888;">
      <div style="font-size:64px">📊</div>
      <div class="dashboard-title" style="text-align:center;">Gladwells</div>
      <p style="font-size:16px; margin-top:12px;">
        {"Connecting to Google Drive…" if folder_configured else
         "Add <code>GDRIVE_FOLDER_ID</code> to your Streamlit secrets, or paste the folder ID in the sidebar."}
      </p>
      <p style="font-size:13px; color:#aaa; max-width:480px; margin:0 auto;">
        <b>Monthly workflow:</b> Drop the new .xlsm file into the shared Drive folder —
        the dashboard will automatically pick it up on the next load or refresh.
      </p>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

try:
    data = parse_workbook(buf.read())
except Exception as e:
    st.error(f"Failed to parse workbook: {e}")
    st.stop()

months   = data["months"]
curr_mth = months[0] if months else "Current"
prev_mth = months[1] if len(months) > 1 else "Prior"
# Prior year label: same month, 12 months back
prev_yr_mth = months[12] if len(months) > 12 else "Prior Year"

# ─── Header ───────────────────────────────────────────────────────────────────
st.markdown(f"""
<div style="display:flex; align-items:center; gap:20px; margin-bottom:16px; padding: 8px 0;">
  <div>
    <div class="dashboard-title">Gladwells Camberwell</div>
    <p class="dashboard-subtitle">{data.get('period_label','Management Accounts')}</p>
  </div>
</div>
""", unsafe_allow_html=True)

# ─── Tabs ─────────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4 = st.tabs(["🏠  Overview", "📈  P&L Detail", "💰  Cash & Balance Sheet", "🛒  Revenue Mix"])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════
with tab1:
    sales_curr   = data.get("ov_sales_curr")
    sales_prev   = data.get("ov_sales_prev_mth")
    sales_prev_yr= data.get("ov_sales_prev_yr")
    gp_curr      = data.get("ov_gp_curr")
    gp_prev      = data.get("ov_gp_prev_mth")
    gp_prev_yr   = data.get("ov_gp_prev_yr")
    gp_pct_c     = data.get("ov_gp_pct_curr")
    gp_pct_p     = data.get("ov_gp_pct_prev_mth")
    gp_pct_py    = data.get("ov_gp_pct_prev_yr")
    np_curr      = data["net_profit"][0]  if data["net_profit"] else None
    np_prev      = data["net_profit"][1]  if len(data["net_profit"]) > 1 else None
    np_prev_yr   = data["net_profit"][12] if len(data["net_profit"]) > 12 else None
    txn_c        = data.get("txn_curr");   txn_p = data.get("txn_prev"); txn_py = data.get("txn_prev_yr")
    asp_c        = data.get("asp_curr");   asp_p = data.get("asp_prev"); asp_py = data.get("asp_prev_yr")
    ytd          = data.get("ov_ytd");     ytd_ly = data.get("ov_ytd_ly")
    cash_c       = data["stat_cash"][-1]  if data.get("stat_cash") else None
    cash_p       = data["stat_cash"][-2]  if data.get("stat_cash") and len(data["stat_cash"]) > 1 else None

    def kpi_card(title, val_str, delta_html_str, extra_class=""):
        return f"""
        <div class="metric-card {extra_class}">
          <div class="metric-title">{title}</div>
          <div class="metric-value">{val_str}</div>
          <div class="metric-delta">{delta_html_str}</div>
        </div>"""

    # ── Row headers + spacer column ──
    spacer, hdr1, hdr2, hdr3, hdr4 = st.columns([1.2, 1, 1, 1, 1])
    spacer.markdown("")
    hdr1.markdown(f'<div class="row-label">Revenue</div>', unsafe_allow_html=True)
    hdr2.markdown(f'<div class="row-label">Gross Profit</div>', unsafe_allow_html=True)
    hdr3.markdown(f'<div class="row-label">GP %</div>', unsafe_allow_html=True)
    hdr4.markdown(f'<div class="row-label">Net Profit</div>', unsafe_allow_html=True)

    # ── Row 1: vs Prior Month ──
    lbl_col, c1, c2, c3, c4 = st.columns([1.2, 1, 1, 1, 1])
    lbl_col.markdown(f'<div style="padding-top:18px; font-size:13px; font-weight:700; color:{PRIMARY};">📅 {curr_mth}</div>', unsafe_allow_html=True)
    c1.markdown(kpi_card(f"vs {prev_mth}", fmt_k(sales_curr), delta_html(sales_curr, sales_prev, label="vs prior month")), unsafe_allow_html=True)
    c2.markdown(kpi_card(f"vs {prev_mth}", fmt_k(gp_curr), delta_html(gp_curr, gp_prev, label="vs prior month")), unsafe_allow_html=True)
    c3.markdown(kpi_card(f"vs {prev_mth}", fmt_pct(gp_pct_c), delta_html(gp_pct_c, gp_pct_p, is_pct=True, label="vs prior month")), unsafe_allow_html=True)
    c4.markdown(kpi_card(f"vs {prev_mth}", fmt_k(np_curr), delta_html(np_curr, np_prev, label="vs prior month")), unsafe_allow_html=True)

    # ── Row 2: vs Prior Year ──
    lbl_col2, d1, d2, d3, d4 = st.columns([1.2, 1, 1, 1, 1])
    lbl_col2.markdown(f'<div style="padding-top:18px; font-size:13px; font-weight:700; color:#777;">📆 vs {prev_yr_mth}</div>', unsafe_allow_html=True)
    d1.markdown(kpi_card(f"vs {prev_yr_mth}", fmt_k(sales_curr), delta_html(sales_curr, sales_prev_yr, label="vs prior year"), "prev-year"), unsafe_allow_html=True)
    d2.markdown(kpi_card(f"vs {prev_yr_mth}", fmt_k(gp_curr), delta_html(gp_curr, gp_prev_yr, label="vs prior year"), "prev-year"), unsafe_allow_html=True)
    d3.markdown(kpi_card(f"vs {prev_yr_mth}", fmt_pct(gp_pct_c), delta_html(gp_pct_c, gp_pct_py, is_pct=True, label="vs prior year"), "prev-year"), unsafe_allow_html=True)
    d4.markdown(kpi_card(f"vs {prev_yr_mth}", fmt_k(np_curr), delta_html(np_curr, np_prev_yr, label="vs prior year"), "prev-year"), unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── KPI row 2: ops metrics ──
    spacer2, hdr5, hdr6, hdr7, hdr8 = st.columns([1.2, 1, 1, 1, 1])
    hdr5.markdown('<div class="row-label">Transactions p.w.</div>', unsafe_allow_html=True)
    hdr6.markdown('<div class="row-label">Avg Spend / Head</div>', unsafe_allow_html=True)
    hdr7.markdown('<div class="row-label">YTD Revenue</div>', unsafe_allow_html=True)
    hdr8.markdown('<div class="row-label">Cash Balance</div>', unsafe_allow_html=True)

    lbl_col3, e1, e2, e3, e4 = st.columns([1.2, 1, 1, 1, 1])
    lbl_col3.markdown(f'<div style="padding-top:18px; font-size:13px; font-weight:700; color:{PRIMARY};">📅 {curr_mth}</div>', unsafe_allow_html=True)

    def ops_delta(curr, prev, fmt_fn, label):
        if curr is None or prev is None: return ""
        diff = curr - prev
        cls = "delta-pos" if diff > 0 else "delta-neg"
        sign = "▲" if diff > 0 else "▼"
        return f'<span class="{cls}">{sign} {fmt_fn(abs(diff))} {label}</span>'

    e1.markdown(kpi_card(f"vs {prev_mth}", f"{txn_c:,.0f}" if txn_c else "—",
        ops_delta(txn_c, txn_p, lambda v: f"{v:,.0f}", "vs prior month")), unsafe_allow_html=True)
    e2.markdown(kpi_card(f"vs {prev_mth}", f"£{asp_c:.2f}" if asp_c else "—",
        ops_delta(asp_c, asp_p, lambda v: f"£{v:.2f}", "vs prior month")), unsafe_allow_html=True)
    e3.markdown(kpi_card("YTD vs LY", fmt_k(ytd), delta_html(ytd, ytd_ly, label="vs last year")), unsafe_allow_html=True)
    e4.markdown(kpi_card(f"vs {prev_mth}", fmt_k(cash_c), delta_html(cash_c, cash_p, label="vs prior month")), unsafe_allow_html=True)

    lbl_col4, f1, f2, f3, f4 = st.columns([1.2, 1, 1, 1, 1])
    lbl_col4.markdown(f'<div style="padding-top:18px; font-size:13px; font-weight:700; color:#777;">📆 vs {prev_yr_mth}</div>', unsafe_allow_html=True)
    f1.markdown(kpi_card(f"vs {prev_yr_mth}", f"{txn_c:,.0f}" if txn_c else "—",
        ops_delta(txn_c, txn_py, lambda v: f"{v:,.0f}", "vs prior year"), "prev-year"), unsafe_allow_html=True)
    f2.markdown(kpi_card(f"vs {prev_yr_mth}", f"£{asp_c:.2f}" if asp_c else "—",
        ops_delta(asp_c, asp_py, lambda v: f"£{v:.2f}", "vs prior year"), "prev-year"), unsafe_allow_html=True)
    f3.markdown(kpi_card("YTD vs LY", fmt_k(ytd), delta_html(ytd, ytd_ly, label="vs last year"), "prev-year"), unsafe_allow_html=True)
    f4.markdown(kpi_card(f"vs {prev_yr_mth}", fmt_k(cash_c), "", "prev-year"), unsafe_allow_html=True)

    # ── 12-month trend chart ──
    st.markdown('<div class="section-header">📅 12-Month Trend</div>', unsafe_allow_html=True)
    legend_hint()
    sm = data.get("stat_months", [])
    st_turn = data.get("stat_turnover", [])
    st_gp   = [g*100 for g in data.get("stat_gp_pct", [])]
    st_cash = data.get("stat_cash", [])

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(go.Bar(
        x=sm, y=st_turn, name="Revenue", marker_color=PRIMARY, opacity=0.85,
        text=[f"£{v/1000:.0f}k" for v in st_turn], textposition="outside", textfont=dict(size=9)
    ), secondary_y=False)
    fig.add_trace(go.Scatter(
        x=sm, y=st_gp, name="GP %", mode="lines+markers",
        line=dict(color=ACCENT, width=2.5), marker=dict(size=7)
    ), secondary_y=True)
    fig.add_trace(go.Scatter(
        x=sm, y=[v/1000 for v in st_cash], name="Cash (£k)", mode="lines+markers",
        line=dict(color=ORANGE, width=2, dash="dot"), marker=dict(size=6)
    ), secondary_y=False)
    t = plotly_theme()
    fig.update_layout(
        **t,
        yaxis=dict(title="£ (000s)", tickformat=",.0f"),
        yaxis2=dict(title="GP %", tickformat=".0f", ticksuffix="%"),
        barmode="group",
    )
    st.plotly_chart(fig, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — P&L DETAIL
# ══════════════════════════════════════════════════════════════════════════════
with tab2:
    st.markdown('<div class="section-header">Profit & Loss — Monthly Waterfall</div>', unsafe_allow_html=True)
    rev = data["revenue"][0]      if data["revenue"]      else 0
    cgs = -abs(data["cogs"][0])   if data["cogs"]         else 0
    gp  = data["gross_profit"][0] if data["gross_profit"] else 0
    stf = -abs(data["staff_costs"][0])  if data["staff_costs"]  else 0
    rnt = -abs(data["rent_util"][0])    if data["rent_util"]    else 0
    mnt = -abs(data["maintenance"][0])  if data["maintenance"]  else 0
    ooh = -abs(data["other_oh"][0])     if data["other_oh"]     else 0
    dep = -abs(data["depreciation"][0]) if data["depreciation"] else 0
    np  = data["net_profit"][0]   if data["net_profit"]   else 0

    fig2 = go.Figure(go.Waterfall(
        orientation="v",
        measure=["absolute","relative","total","relative","relative","relative","relative","relative","total"],
        x=["Revenue","Cost of Sales","Gross Profit","Staff","Rent & Utilities","Maintenance","Other Overheads","Depreciation","Net Profit"],
        y=[rev, cgs, gp, stf, rnt, mnt, ooh, dep, np],
        connector=dict(line=dict(color="#ccc", width=1)),
        increasing=dict(marker_color=PRIMARY),
        decreasing=dict(marker_color=ORANGE),
        totals=dict(marker_color=ACCENT),
        text=[fmt_k(abs(v)) for v in [rev, cgs, gp, stf, rnt, mnt, ooh, dep, np]],
        textposition="outside",
    ))
    fig2.update_layout(**plotly_theme(), yaxis=dict(tickformat=",.0f", title="£"), showlegend=False)
    st.plotly_chart(fig2, use_container_width=True)

    # ── Monthly P&L table with highlighted Net Profit ──
    st.markdown('<div class="section-header">Monthly P&L Summary (£000s)</div>', unsafe_allow_html=True)
    n = min(len(months), 13)
    pnl_rows = {
        "Revenue":          data["revenue"][:n],
        "Cost of Sales":    [-abs(v) for v in data["cogs"][:n]],
        "Gross Profit":     data["gross_profit"][:n],
        "GP %":             [f"{g/r*100:.1f}%" if r else "—" for g, r in zip(data["gross_profit"][:n], data["revenue"][:n])],
        "Staff Costs":      [-abs(v) for v in data["staff_costs"][:n]],
        "Rent & Utilities": [-abs(v) for v in data["rent_util"][:n]],
        "Maintenance":      [-abs(v) for v in data["maintenance"][:n]],
        "Other Overheads":  [-abs(v) for v in data["other_oh"][:n]],
        "Total Overheads":  [-abs(v) for v in data["total_opex"][:n]],
        "Operating Profit": data["op_profit"][:n],
        "Depreciation":     [-abs(v) for v in data["depreciation"][:n]],
        "Net Profit":       data["net_profit"][:n],
    }
    col_headers = months[:n]
    table_data = {}
    for k, vals in pnl_rows.items():
        row_vals = []
        for v in vals:
            if isinstance(v, str): row_vals.append(v)
            elif isinstance(v, (float, int)): row_vals.append(f"£{v/1000:,.1f}k")
            else: row_vals.append("—")
        table_data[k] = row_vals

    df_pnl = pd.DataFrame(table_data, index=col_headers).T

    # Style: bold + highlight Net Profit row
    def style_pnl(df):
        styles = pd.DataFrame("", index=df.index, columns=df.columns)
        if "Net Profit" in df.index:
            styles.loc["Net Profit"] = "font-weight: bold; background-color: #e8f5e9; color: #1a2e1a;"
        if "Gross Profit" in df.index:
            styles.loc["Gross Profit"] = "font-weight: 600; background-color: #f0f7ee;"
        return styles

    styled_pnl = df_pnl.style.apply(style_pnl, axis=None)
    st.dataframe(styled_pnl, use_container_width=True, height=430)

    # ── Overhead breakdown ──
    st.markdown('<div class="section-header">Overhead Breakdown — Last 12 Months</div>', unsafe_allow_html=True)
    legend_hint()
    sm2 = list(reversed(months[:12]))
    fig3 = go.Figure()
    oh_items = {
        "Staff":           list(reversed(data["staff_costs"][:12])),
        "Rent & Utilities":list(reversed(data["rent_util"][:12])),
        "Maintenance":     list(reversed(data["maintenance"][:12])),
        "Other":           list(reversed(data["other_oh"][:12])),
    }
    for (lbl, vals), col in zip(oh_items.items(), [PRIMARY, ACCENT, ORANGE, "#B5D4A8"]):
        fig3.add_trace(go.Bar(x=sm2, y=vals, name=lbl, marker_color=col))
    fig3.update_layout(**plotly_theme(), barmode="stack", yaxis=dict(title="£", tickformat=",.0f"))
    st.plotly_chart(fig3, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — CASH & BALANCE SHEET
# ══════════════════════════════════════════════════════════════════════════════
with tab3:
    st.markdown('<div class="section-header">Cash Position — 12 Months</div>', unsafe_allow_html=True)
    legend_hint()
    sm3       = data.get("stat_months", [])
    cash_vals = data.get("stat_cash", [])
    mcf = [cash_vals[i] - cash_vals[i-1] if i > 0 else 0 for i in range(len(cash_vals))]

    fig4 = make_subplots(specs=[[{"secondary_y": True}]])
    fig4.add_trace(go.Bar(
        x=sm3, y=mcf, name="Monthly Cash Flow",
        marker_color=[PRIMARY if v >= 0 else ORANGE for v in mcf], opacity=0.75,
    ), secondary_y=False)
    fig4.add_trace(go.Scatter(
        x=sm3, y=cash_vals, name="Cash Balance", mode="lines+markers",
        line=dict(color=ACCENT, width=3), marker=dict(size=8, color=ACCENT),
        fill="tozeroy", fillcolor="rgba(151,188,98,0.12)"
    ), secondary_y=True)
    fig4.update_layout(
        **plotly_theme(),
        yaxis=dict(title="Monthly Cash Flow (£)", tickformat=",.0f"),
        yaxis2=dict(title="Cash Balance (£)", tickformat=",.0f"),
    )
    st.plotly_chart(fig4, use_container_width=True)

    st.markdown('<div class="section-header">Balance Sheet Snapshot</div>', unsafe_allow_html=True)
    bca   = data.get("bs_cash_bal",   [None])[0]
    bca_p = data.get("bs_cash_bal",   [None,None])[1] if len(data.get("bs_cash_bal",[])) > 1 else None
    bcu   = data.get("bs_current",    [None])[0]
    bcl   = data.get("bs_curr_liab",  [None])[0]
    bna   = data.get("bs_net_assets", [None])[0]

    cc1, cc2, cc3, cc4 = st.columns(4)
    for col, title, curr, prev in [
        (cc1, "Cash & Bank",          bca,  bca_p),
        (cc2, "Net Assets",           bna,  None),
        (cc3, "Current Assets",       bcu,  None),
        (cc4, "Current Liabilities",  bcl,  None),
    ]:
        d = delta_html(curr, prev) if prev is not None else ""
        col.markdown(f"""
        <div class="metric-card">
          <div class="metric-title">{title}</div>
          <div class="metric-value">{fmt_k(curr)}</div>
          <div class="metric-delta">{d}</div>
        </div>""", unsafe_allow_html=True)

    legend_hint()
    bs_m    = data.get("bs_months", [])
    bs_ca   = data.get("bs_current", [])
    bs_cl   = [abs(v) for v in data.get("bs_curr_liab", [])]
    bs_na   = data.get("bs_net_assets", [])
    bs_cash = data.get("bs_cash_bal", [])
    if bs_m:
        n_bs = min(len(bs_m), len(bs_ca), len(bs_cl), len(bs_na))
        fig5 = go.Figure()
        fig5.add_trace(go.Scatter(x=bs_m[:n_bs], y=bs_ca[:n_bs],   name="Current Assets",       mode="lines+markers", line=dict(color=PRIMARY, width=2)))
        fig5.add_trace(go.Scatter(x=bs_m[:n_bs], y=bs_cl[:n_bs],   name="Current Liabilities",  mode="lines+markers", line=dict(color=ORANGE,  width=2, dash="dash")))
        fig5.add_trace(go.Scatter(x=bs_m[:n_bs], y=bs_cash[:n_bs], name="Cash",                 mode="lines+markers", line=dict(color=ACCENT,  width=2)))
        fig5.add_trace(go.Scatter(x=bs_m[:n_bs], y=bs_na[:n_bs],   name="Net Assets",           mode="lines+markers", line=dict(color="#7A5C58",width=2, dash="dot")))
        fig5.update_layout(**plotly_theme(), yaxis=dict(title="£", tickformat=",.0f"))
        st.plotly_chart(fig5, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — REVENUE MIX
# ══════════════════════════════════════════════════════════════════════════════
with tab4:
    st.markdown('<div class="section-header">Revenue Mix — Current Month</div>', unsafe_allow_html=True)
    rb         = data.get("revenue_breakdown", {})
    labels_pie = [k for k, v in rb.items() if v and v[0] and v[0] > 0]
    vals_pie   = [rb[k][0] for k in labels_pie]
    colour_seq = [PRIMARY, ACCENT, ORANGE, "#B5D4A8", "#7A5C58", "#D4A853", "#5C8A7A", "#C47A5A", "#6B8CBF", "#A87A5C"]

    col_a, col_b = st.columns([1, 1])
    with col_a:
        legend_hint()
        fig6 = go.Figure(go.Pie(
            labels=labels_pie, values=vals_pie,
            hole=0.5, marker=dict(colors=colour_seq),
            textinfo="label+percent", textfont=dict(size=11),
        ))
        fig6.add_annotation(text=f"<b>Total</b><br>{fmt_k(sum(vals_pie))}", x=0.5, y=0.5, font_size=13, showarrow=False)
        t6 = plotly_theme()
        t6["margin"] = dict(l=20, r=20, t=55, b=20)
        fig6.update_layout(**t6, showlegend=True)
        st.plotly_chart(fig6, use_container_width=True)

    with col_b:
        df_rb = pd.DataFrame({
            "Category":   labels_pie,
            curr_mth:     [fmt_k(rb[k][0]) for k in labels_pie],
            prev_mth:     [fmt_k(rb[k][1]) if len(rb[k]) > 1 else "—" for k in labels_pie],
            "Change":     [
                f"{'▲' if rb[k][0] > rb[k][1] else '▼'} {fmt_k(abs(rb[k][0]-rb[k][1]))}"
                if len(rb[k]) > 1 and rb[k][1] else "—"
                for k in labels_pie
            ]
        })
        st.dataframe(df_rb, use_container_width=True, hide_index=True, height=360)

    st.markdown('<div class="section-header">Revenue Trend by Category — Last 12 Months</div>', unsafe_allow_html=True)
    legend_hint()
    top_cats = sorted(labels_pie, key=lambda k: rb[k][0], reverse=True)[:5]
    sm4 = list(reversed(months[:12]))
    fig7 = go.Figure()
    for cat, col in zip(top_cats, colour_seq):
        vals = list(reversed(rb[cat][:12]))
        fig7.add_trace(go.Bar(x=sm4, y=vals, name=cat, marker_color=col))
    fig7.update_layout(**plotly_theme(), barmode="stack", yaxis=dict(title="£", tickformat=",.0f"))
    st.plotly_chart(fig7, use_container_width=True)

# ─── Footer ───────────────────────────────────────────────────────────────────
st.markdown(f"""
<div style="text-align:center; color:#bbb; font-size:12px; margin-top:40px; padding:20px;">
  Gladwells Management Accounts Dashboard · Auto-refreshes from Google Drive every hour ·
  Last loaded: {datetime.now().strftime("%d %b %Y %H:%M")}
</div>""", unsafe_allow_html=True)
